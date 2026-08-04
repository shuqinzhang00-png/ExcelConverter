from __future__ import annotations

import io
import re
import unicodedata
from collections import Counter, defaultdict
from typing import Any, BinaryIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =========================================================
# General utility functions
# =========================================================

def clean_value(value: Any) -> str:
    """
    Convert a value to a trimmed string.

    None, pandas NA, and NaN values become an empty string.
    """

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    return str(value).strip()


def normalize_column_name(value: Any) -> str:
    """
    Normalize an Excel column name.

    Examples:
        " SP "       -> "sp"
        "Reference"  -> "reference"
        "Label Name" -> "label name"
    """

    text = clean_value(value).lower()

    # Replace repeated spaces with one space.
    return re.sub(r"\s+", " ", text)


def normalize_reference(value: Any) -> str:
    """
    Remove an existing generated reference suffix.

    Examples:
        y050          -> y050
        y050-01/08    -> y050
        ABC-001/125   -> ABC
    """

    reference = clean_value(value)

    reference = re.sub(
        r"-\d+/\d+$",
        "",
        reference,
    )

    return reference.strip()


# =========================================================
# Label processing
# =========================================================

def split_labels(value: Any) -> list[str]:
    """
    Split a label cell using English or Chinese commas.

    Supported separators:
        ,
        ，

    Examples:
        "A, B, C"   -> ["A", "B", "C"]
        "甲，乙，丙" -> ["甲", "乙", "丙"]
    """

    text = clean_value(value)

    if not text:
        return []

    parts = re.split(
        r"\s*[,，]\s*",
        text,
    )

    return [
        part.strip()
        for part in parts
        if part.strip()
    ]


def calculate_visual_width(value: Any) -> int:
    """
    Calculate approximate visual text width.

    Chinese and other full-width characters count as 2.
    Most English characters count as 1.
    """

    text = clean_value(value)
    width = 0

    for character in text:
        east_asian_width = unicodedata.east_asian_width(
            character
        )

        if east_asian_width in {"W", "F"}:
            width += 2
        else:
            width += 1

    return width


def determine_label_size(
    label: str,
    large_width: int,
    medium_width: int,
) -> tuple[str, int]:
    """
    Classify a label as large, medium, or small.

    The returned width is the calculated visual width.
    """

    visual_width = calculate_visual_width(label)

    if visual_width <= large_width:
        return "large", visual_width

    if visual_width <= medium_width:
        return "medium", visual_width

    return "small", visual_width


# =========================================================
# YPW SP-name extraction
# =========================================================

def normalize_parentheses(value: Any) -> str:
    """
    Convert Chinese parentheses to English parentheses.
    """

    return (
        clean_value(value)
        .replace("（", "(")
        .replace("）", ")")
    )
def clean_ypw_name(value: Any) -> str:
    """
    Clean a name extracted from a YPW marker.

    Supports both:
        陽上
        阳上
    """

    name = clean_value(value)

    # Remove Traditional or Simplified YPW marker.
    name = re.sub(
        r"^(?:陽上|阳上)\s*",
        "",
        name,
    ).strip()

    # Remove surrounding parentheses and whitespace.
    name = name.strip("()（） \t\r\n")

    return name
    
def extract_names_from_parentheses(
    value: Any,
) -> list[str]:
    """
    Extract names from all parenthesized sections.

    Examples:
        吳氏歷代祖先(阳上吳大郎)
            -> ["吳大郎"]

        謝欣娟的冤親債主（陽上謝欣娟）
            -> ["謝欣娟"]

        (陽上林玉瑩)
            -> ["林玉瑩"]

        （陳美玲）
            -> ["陳美玲"]
    """

    text = normalize_parentheses(value)

    parenthesized_values = re.findall(
        r"\(([^()]*)\)",
        text,
    )

    names: list[str] = []

    for parenthesized_value in parenthesized_values:
        name = clean_ypw_name(
            parenthesized_value
        )

        if name:
            names.append(name)

    return names
def remove_parenthesized_names(
    value: Any,
) -> str:
    """
    Remove parenthesized name sections from a label.

    Examples:
        吳氏歷代祖先(阳上吳大郎)
            -> 吳氏歷代祖先

        謝欣娟的冤親債主（陽上謝欣娟）
            -> 謝欣娟的冤親債主

        (陽上林玉瑩)
            -> ""
    """

    text = normalize_parentheses(value)

    cleaned_label = re.sub(
        r"\s*\([^()]*\)\s*",
        "",
        text,
    )

    # Remove unnecessary spaces around the result.
    cleaned_label = re.sub(
        r"\s+",
        " ",
        cleaned_label,
    )

    return cleaned_label.strip()


def unique_values(values: list[str]) -> list[str]:
    """
    Remove duplicates while preserving original order.
    """

    result: list[str] = []
    seen: set[str] = set()

    for value in values:
        cleaned_value = clean_value(value)

        if cleaned_value and cleaned_value not in seen:
            seen.add(cleaned_value)
            result.append(cleaned_value)

    return result

def process_ypw_labels(
    source_labels: list[str],
    original_sp: str,
) -> list[dict[str, str]]:
    """
    Convert labels for one YPW source row.

    Rules
    -----
    1. An ordinary label containing a parenthesized name:

           吳氏歷代祖先(阳上吳大郎)

       becomes:

           sp    = 吳大郎
           label = 吳氏歷代祖先

    2. Standalone parenthesized names following an ordinary
       label are combined with that ordinary label:

           普通名稱, (陽上林玉瑩), （陳美玲）

       becomes one output record:

           sp    = 林玉瑩 陳美玲
           label = 普通名稱

    3. A label without a parenthesized name keeps the
       original SP.
    """

    processed_records: list[dict[str, str]] = []

    current_record: dict[str, Any] | None = None

    for source_label in source_labels:
        source_label = clean_value(source_label)

        if not source_label:
            continue

        extracted_names = extract_names_from_parentheses(
            source_label
        )

        cleaned_label = remove_parenthesized_names(
            source_label
        )

        # -------------------------------------------------
        # Standalone parenthesized item:
        #
        # (陽上林玉瑩)
        # （陳美玲）
        #
        # cleaned_label is empty, so attach its name to
        # the preceding ordinary label.
        # -------------------------------------------------
        if not cleaned_label and extracted_names:
            if current_record is not None:
                current_record["names"].extend(
                    extracted_names
                )
            else:
                # There is no preceding label. Preserve it
                # temporarily so it can be reported.
                current_record = {
                    "label": "",
                    "names": extracted_names.copy(),
                }

            continue

        # -------------------------------------------------
        # Ordinary label: finish the previous record first.
        # -------------------------------------------------
        if current_record is not None:
            names = unique_values(
                current_record["names"]
            )

            processed_records.append(
                {
                    "sp": (
                        " ".join(names)
                        if names
                        else clean_value(original_sp)
                    ),
                    "label": clean_value(
                        current_record["label"]
                    ),
                }
            )

        current_record = {
            "label": cleaned_label,
            "names": extracted_names.copy(),
        }

    # Add the final record.
    if current_record is not None:
        names = unique_values(
            current_record["names"]
        )

        processed_records.append(
            {
                "sp": (
                    " ".join(names)
                    if names
                    else clean_value(original_sp)
                ),
                "label": clean_value(
                    current_record["label"]
                ),
            }
        )

    # Remove records that have no actual output label.
    return [
        record
        for record in processed_records
        if record["label"]
    ]


def clean_extracted_name(value: Any) -> str:
    """
    Clean a name extracted from a YPW label.
    """

    name = clean_value(value)

    # Remove the YPW marker from the beginning.
    name = re.sub(
        r"^陽上\s*",
        "",
        name,
    )

    # Remove surrounding parentheses and spaces.
    name = name.strip("()（） \t\r\n")

    # Stop at common separators.
    name = re.split(
        r"[,，、;；:/]",
        name,
        maxsplit=1,
    )[0]

    return name.strip()


def extract_ypw_sp(
    label: str,
    original_sp: str,
) -> str:
    """
    Extract the SP name for YPW conversion.

    Examples:
        (陽上林玉瑩)     -> 林玉瑩
        （陽上林玉瑩）   -> 林玉瑩
        陽上林玉瑩       -> 林玉瑩
        (林玉瑩)         -> 林玉瑩
        （林玉瑩）       -> 林玉瑩

    When no name can be extracted, the original SP is used.
    """

    text = normalize_parentheses(label)
    original_sp = clean_value(original_sp)

    if not text:
        return original_sp

    # -----------------------------------------------------
    # Priority 1:
    # Extract text following 陽上 inside parentheses.
    #
    # Example:
    #   超薦(陽上林玉瑩)
    #   (陽上林玉瑩)
    # -----------------------------------------------------
    match = re.search(
        r"\(\s*陽上\s*([^()]+?)\s*\)",
        text,
    )

    if match:
        name = clean_extracted_name(
            match.group(1)
        )

        if name:
            return name

    # -----------------------------------------------------
    # Priority 2:
    # Extract text following 陽上 without requiring
    # parentheses.
    #
    # Example:
    #   陽上林玉瑩
    # -----------------------------------------------------
    match = re.search(
        r"陽上\s*([^()，,、;；:/]+)",
        text,
    )

    if match:
        name = clean_extracted_name(
            match.group(1)
        )

        if name:
            return name

    # -----------------------------------------------------
    # Priority 3:
    # Extract ordinary parenthesized content.
    #
    # Example:
    #   (林玉瑩)
    # -----------------------------------------------------
    matches = re.findall(
        r"\(([^()]+)\)",
        text,
    )

    for parenthesized_text in matches:
        name = clean_extracted_name(
            parenthesized_text
        )

        # Do not return only the marker itself.
        if name and name != "陽上":
            return name

    return original_sp


# =========================================================
# Input validation
# =========================================================

def normalize_source_columns(
    source_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Normalize source column names and accept common aliases.
    """

    source_df = source_df.copy()

    source_df.columns = [
        normalize_column_name(column)
        for column in source_df.columns
    ]

    column_aliases = {
        # SP aliases
        "name": "sp",
        "sp name": "sp",
        "sponsor": "sp",

        # Reference aliases
        "reference": "ref",
        "reference no": "ref",
        "reference number": "ref",
        "ref no": "ref",

        # Label aliases
        "labels": "label",
        "label name": "label",
        "label names": "label",
    }

    source_df = source_df.rename(
        columns=column_aliases
    )

    return source_df


def validate_columns(
    source_df: pd.DataFrame,
) -> None:
    """
    Confirm that all required input columns exist.
    """

    required_columns = {
        "sp",
        "ref",
        "label",
    }

    available_columns = set(
        source_df.columns
    )

    missing_columns = (
        required_columns
        - available_columns
    )

    if missing_columns:
        missing_text = ", ".join(
            sorted(missing_columns)
        )

        available_text = ", ".join(
            str(column)
            for column in source_df.columns
        )

        raise ValueError(
            f"Missing required column(s): {missing_text}. "
            f"Available columns are: {available_text}"
        )


# =========================================================
# Reference generation
# =========================================================

def generate_reference(
    base_reference: str,
    record_number: int,
    total_records: int,
) -> str:
    """
    Generate an automatically numbered reference.

    Examples:
        y050, 1, 8   -> y050-01/08
        y050, 8, 8   -> y050-08/08
        y100, 1, 125 -> y100-001/125
    """

    base_reference = normalize_reference(
        base_reference
    )

    if not base_reference:
        return ""

    if record_number < 1 or total_records < 1:
        return base_reference

    number_width = max(
        2,
        len(str(total_records)),
    )

    return (
        f"{base_reference}-"
        f"{record_number:0{number_width}d}/"
        f"{total_records:0{number_width}d}"
    )


# =========================================================
# Expand source data
# =========================================================

def expand_source_data(
    source_df: pd.DataFrame,
    large_width: int,
    medium_width: int,
    conversion_type: str = "HPW",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Expand the source Excel records.

    HPW:
        Split comma-separated labels into separate records.
        Keep the original SP and original label text.

    YPW:
        Extract names from parentheses, remove the
        parentheses from the label, and use the extracted
        names as SP.
    """

    conversion_type = clean_value(
        conversion_type
    ).upper()

    if conversion_type not in {"HPW", "YPW"}:
        raise ValueError(
            "Conversion type must be HPW or YPW."
        )

    if large_width < 1:
        raise ValueError(
            "Large-font maximum width must be at least 1."
        )

    if medium_width <= large_width:
        raise ValueError(
            "Medium-font maximum width must be greater "
            "than large-font maximum width."
        )

    expanded_rows: list[dict[str, Any]] = []
    validation_rows: list[dict[str, Any]] = []

    # =====================================================
    # First pass: split and process all source rows
    # =====================================================
    for source_index, source_row in source_df.iterrows():
        excel_row = source_index + 2

        original_sp = clean_value(
            source_row.get("sp", "")
        )

        base_reference = normalize_reference(
            source_row.get("ref", "")
        )

        source_labels = split_labels(
            source_row.get("label", "")
        )

        if not base_reference:
            validation_rows.append(
                {
                    "severity": "Warning",
                    "source_row": excel_row,
                    "sp": original_sp,
                    "ref": "",
                    "label": "",
                    "message": "The reference is empty.",
                }
            )

        if not source_labels:
            validation_rows.append(
                {
                    "severity": "Warning",
                    "source_row": excel_row,
                    "sp": original_sp,
                    "ref": base_reference,
                    "label": "",
                    "message": "No valid labels were found.",
                }
            )
            continue

        # -------------------------------------------------
        # Create output records according to type.
        # -------------------------------------------------
        if conversion_type == "YPW":
            processed_records = process_ypw_labels(
                source_labels=source_labels,
                original_sp=original_sp,
            )
        else:
            processed_records = [
                {
                    "sp": original_sp,
                    "label": label,
                }
                for label in source_labels
            ]

        if not processed_records:
            validation_rows.append(
                {
                    "severity": "Warning",
                    "source_row": excel_row,
                    "sp": original_sp,
                    "ref": base_reference,
                    "label": clean_value(
                        source_row.get("label", "")
                    ),
                    "message": (
                        "No output label remained after "
                        "YPW name extraction."
                    ),
                }
            )
            continue

        # -------------------------------------------------
        # Add each processed record.
        # -------------------------------------------------
        for label_order, processed_record in enumerate(
            processed_records,
            start=1,
        ):
            output_sp = clean_value(
                processed_record["sp"]
            )

            output_label = clean_value(
                processed_record["label"]
            )

            # Important:
            # Calculate font size from the cleaned output
            # label, not from the original text containing
            # parentheses.
            label_size, visual_width = (
                determine_label_size(
                    label=output_label,
                    large_width=large_width,
                    medium_width=medium_width,
                )
            )

            expanded_rows.append(
                {
                    "source_order": source_index,
                    "source_row": excel_row,
                    "label_order": label_order,
                    "conversion_type": conversion_type,
                    "original_sp": original_sp,
                    "sp": output_sp,
                    "base_ref": base_reference,
                    "label": output_label,

                    # Only one font-size field contains text.
                    "label_large": (
                        output_label
                        if label_size == "large"
                        else ""
                    ),
                    "label_medium": (
                        output_label
                        if label_size == "medium"
                        else ""
                    ),
                    "label_small": (
                        output_label
                        if label_size == "small"
                        else ""
                    ),

                    "label_size": label_size,
                    "label_visual_width": visual_width,
                }
            )

    if not expanded_rows:
        raise ValueError(
            "No output records were created."
        )

    # =====================================================
    # Count records for every base reference
    # =====================================================
    reference_totals = Counter(
        row["base_ref"]
        for row in expanded_rows
        if row["base_ref"]
    )

    reference_sequence: defaultdict[str, int] = (
        defaultdict(int)
    )

    # =====================================================
    # Assign references
    # =====================================================
    for row in expanded_rows:
        base_reference = row["base_ref"]

        if not base_reference:
            row["label_number"] = ""
            row["total_labels"] = ""
            row["ref"] = ""
            continue

        reference_sequence[base_reference] += 1

        record_number = reference_sequence[
            base_reference
        ]

        total_records = reference_totals[
            base_reference
        ]

        row["label_number"] = record_number
        row["total_labels"] = total_records

        row["ref"] = generate_reference(
            base_reference=base_reference,
            record_number=record_number,
            total_records=total_records,
        )

    # =====================================================
    # Build DataFrames
    # =====================================================
    expanded_df = pd.DataFrame(
        expanded_rows
    )

    expanded_df = expanded_df.sort_values(
        by=[
            "source_order",
            "label_order",
        ],
        kind="stable",
    ).reset_index(drop=True)

    expanded_df = expanded_df[
        [
            "source_row",
            "conversion_type",
            "original_sp",
            "sp",
            "base_ref",
            "ref",
            "label_number",
            "total_labels",
            "label",
            "label_large",
            "label_medium",
            "label_small",
            "label_size",
            "label_visual_width",
        ]
    ]

    validation_df = pd.DataFrame(
        validation_rows,
        columns=[
            "severity",
            "source_row",
            "sp",
            "ref",
            "label",
            "message",
        ],
    )

    return expanded_df, validation_df


# =========================================================
# Create configurable Mail Merge rows
# =========================================================

def create_mail_merge_data(
    expanded_df: pd.DataFrame,
    group_size: int,
) -> pd.DataFrame:
    """
    Combine expanded records into configurable output rows.

    For group_size=4, columns include:
        sp1, ref1, label1, ...
        sp2, ref2, label2, ...
        sp3, ref3, label3, ...
        sp4, ref4, label4, ...

    The large, medium, and small font fields are also
    created for every position.
    """

    group_size = int(group_size)

    if group_size < 1:
        raise ValueError(
            "Records per output row must be at least 1."
        )

    if group_size > 50:
        raise ValueError(
            "Records per output row cannot exceed 50."
        )

    output_rows: list[dict[str, Any]] = []

    for start_index in range(
        0,
        len(expanded_df),
        group_size,
    ):
        group = expanded_df.iloc[
            start_index:start_index + group_size
        ]

        output_row: dict[str, Any] = {}

        for position in range(group_size):
            number = position + 1

            if position < len(group):
                record = group.iloc[position]

                output_row[
                    f"sp{number}"
                ] = clean_value(
                    record.get("sp", "")
                )

                output_row[
                    f"ref{number}"
                ] = clean_value(
                    record.get("ref", "")
                )

                output_row[
                    f"label{number}"
                ] = clean_value(
                    record.get("label", "")
                )

                output_row[
                    f"label{number}_large"
                ] = clean_value(
                    record.get("label_large", "")
                )

                output_row[
                    f"label{number}_medium"
                ] = clean_value(
                    record.get("label_medium", "")
                )

                output_row[
                    f"label{number}_small"
                ] = clean_value(
                    record.get("label_small", "")
                )

                output_row[
                    f"label{number}_size"
                ] = clean_value(
                    record.get("label_size", "")
                )

                output_row[
                    f"label{number}_visual_width"
                ] = record.get(
                    "label_visual_width",
                    "",
                )

                output_row[
                    f"original_sp{number}"
                ] = clean_value(
                    record.get("original_sp", "")
                )

            else:
                # Fill unused positions in the final row.
                output_row[f"sp{number}"] = ""
                output_row[f"ref{number}"] = ""
                output_row[f"label{number}"] = ""
                output_row[
                    f"label{number}_large"
                ] = ""
                output_row[
                    f"label{number}_medium"
                ] = ""
                output_row[
                    f"label{number}_small"
                ] = ""
                output_row[
                    f"label{number}_size"
                ] = ""
                output_row[
                    f"label{number}_visual_width"
                ] = ""
                output_row[
                    f"original_sp{number}"
                ] = ""

        output_rows.append(output_row)

    return pd.DataFrame(output_rows)


# =========================================================
# Workbook formatting
# =========================================================

def calculate_excel_column_width(
    worksheet,
    column_index: int,
    maximum_width: int = 50,
) -> int:
    """
    Calculate a reasonable Excel column width.
    """

    maximum_length = 0

    for row_index in range(
        1,
        worksheet.max_row + 1,
    ):
        value = worksheet.cell(
            row=row_index,
            column=column_index,
        ).value

        text = clean_value(value)

        if not text:
            continue

        visual_length = calculate_visual_width(
            text
        )

        maximum_length = max(
            maximum_length,
            visual_length,
        )

    return min(
        max(maximum_length + 2, 10),
        maximum_width,
    )


def format_excel_workbook(
    source_buffer: io.BytesIO,
) -> io.BytesIO:
    """
    Apply formatting to all generated worksheets.
    """

    source_buffer.seek(0)

    workbook = load_workbook(
        source_buffer
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    information_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        if worksheet.max_row >= 1:
            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

        worksheet.row_dimensions[1].height = 24

        # Header formatting.
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # Data formatting.
        for row in worksheet.iter_rows(
            min_row=2,
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        # Dynamic column widths.
        for column_index in range(
            1,
            worksheet.max_column + 1,
        ):
            column_letter = get_column_letter(
                column_index
            )

            worksheet.column_dimensions[
                column_letter
            ].width = calculate_excel_column_width(
                worksheet=worksheet,
                column_index=column_index,
            )

        # Make the first data row slightly distinct for
        # worksheets that contain generated information.
        if (
            worksheet.title == "Configuration"
            and worksheet.max_row >= 2
        ):
            for cell in worksheet[2]:
                cell.fill = information_fill

    output_buffer = io.BytesIO()

    workbook.save(
        output_buffer
    )

    output_buffer.seek(0)

    return output_buffer


# =========================================================
# Main conversion function
# =========================================================

def convert_excel(
    source_file: BinaryIO,
    group_size: int = 4,
    large_width: int = 16,
    medium_width: int = 28,
    conversion_type: str = "HPW",
) -> dict[str, Any]:
    """
    Convert an uploaded source Excel workbook into a
    Word Mail Merge workbook.

    Required source columns:
        sp
        ref
        label
    """

    group_size = int(group_size)
    large_width = int(large_width)
    medium_width = int(medium_width)

    conversion_type = clean_value(
        conversion_type
    ).upper()

    if group_size < 1:
        raise ValueError(
            "Records per output row must be at least 1."
        )

    if large_width < 1:
        raise ValueError(
            "Large-font maximum width must be at least 1."
        )

    if medium_width <= large_width:
        raise ValueError(
            "Medium-font maximum width must be greater "
            "than large-font maximum width."
        )

    if conversion_type not in {
        "HPW",
        "YPW",
    }:
        raise ValueError(
            "Conversion type must be HPW or YPW."
        )

    # Reset the uploaded file before reading it.
    source_file.seek(0)

    try:
        source_df = pd.read_excel(
            source_file,
            engine="openpyxl",
        )
    except Exception as error:
        raise ValueError(
            "The uploaded Excel workbook could not be "
            f"read: {error}"
        ) from error

    if source_df.empty:
        raise ValueError(
            "The uploaded Excel workbook contains no rows."
        )

    source_df = normalize_source_columns(
        source_df
    )

    validate_columns(
        source_df
    )

    # Convert only the required input fields to clean strings.
    for column in [
        "sp",
        "ref",
        "label",
    ]:
        source_df[column] = source_df[
            column
        ].map(clean_value)

    expanded_df, validation_df = (
        expand_source_data(
            source_df=source_df,
            large_width=large_width,
            medium_width=medium_width,
            conversion_type=conversion_type,
        )
    )

    mail_merge_df = create_mail_merge_data(
        expanded_df=expanded_df,
        group_size=group_size,
    )

    configuration_df = pd.DataFrame(
        [
            {
                "conversion_type": conversion_type,
                "records_per_output_row": group_size,
                "large_font_maximum_width": large_width,
                "medium_font_maximum_width": medium_width,
                "original_rows": len(source_df),
                "expanded_records": len(expanded_df),
                "mail_merge_rows": len(mail_merge_df),
            }
        ]
    )

    raw_buffer = io.BytesIO()

    with pd.ExcelWriter(
        raw_buffer,
        engine="openpyxl",
    ) as writer:
        mail_merge_df.to_excel(
            writer,
            sheet_name="MailMerge",
            index=False,
        )

        expanded_df.to_excel(
            writer,
            sheet_name="ExpandedData",
            index=False,
        )

        source_df.to_excel(
            writer,
            sheet_name="OriginalData",
            index=False,
        )

        validation_df.to_excel(
            writer,
            sheet_name="Validation",
            index=False,
        )

        configuration_df.to_excel(
            writer,
            sheet_name="Configuration",
            index=False,
        )

    formatted_buffer = format_excel_workbook(
        raw_buffer
    )

    return {
        "output": formatted_buffer,
        "original": source_df,
        "expanded": expanded_df,
        "mail_merge": mail_merge_df,
        "validation": validation_df,
        "configuration": configuration_df,
    }
