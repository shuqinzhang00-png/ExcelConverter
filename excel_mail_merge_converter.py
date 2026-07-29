from __future__ import annotations

import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# Configuration
# ============================================================

INPUT_FILE = Path("input.xlsx")
OUTPUT_FILE = Path("mail_merge_output.xlsx")

# Number of labels placed on one Word Mail Merge record.
GROUP_SIZE = 4

# Word font sizes.
LARGE_FONT_SIZE = 14
MEDIUM_FONT_SIZE = 11
SMALL_FONT_SIZE = 9

# Approximate visual-width limits.
#
# Chinese characters normally occupy more horizontal space than
# English letters, so the program calculates visual width rather
# than only using len(text).
LARGE_LABEL_MAX_WIDTH = 16
MEDIUM_LABEL_MAX_WIDTH = 28

# Characters accepted as label separators.
#
# Supports:
#   English comma: ,
#   Chinese comma: ，
LABEL_SEPARATOR_PATTERN = r"\s*[,，]\s*"


# ============================================================
# Reference rules
# ============================================================

@dataclass(frozen=True)
class ReferenceRule:
    """
    Describes how reference suffixes are generated.

    Example:
        base_ref = y015
        start = 4
        step = 4
        total = 13

    Generated references:
        y015-04/13
        y015-08/13
        y015-12/13
    """

    start: int
    step: int
    total: int


# These rules are based on the examples supplied.
#
# Add or modify rules when new reference numbers appear.
REF_RULES: dict[str, ReferenceRule] = {
    "y007": ReferenceRule(start=1, step=1, total=1),
    "y015": ReferenceRule(start=4, step=4, total=13),
    "y016": ReferenceRule(start=3, step=1, total=5),
    "y017": ReferenceRule(start=2, step=1, total=5),
    "y018": ReferenceRule(start=1, step=1, total=4),
    "y020": ReferenceRule(start=1, step=4, total=12),
    "y025": ReferenceRule(start=1, step=1, total=1),
    "y029": ReferenceRule(start=1, step=1, total=4),
    "y037": ReferenceRule(start=3, step=1, total=3),
    "y040": ReferenceRule(start=2, step=4, total=10),
    "y046": ReferenceRule(start=4, step=1, total=5),
    "y047": ReferenceRule(start=3, step=1, total=6),
    "y048": ReferenceRule(start=1, step=4, total=8),
    "y049": ReferenceRule(start=1, step=4, total=10),
}


# ============================================================
# General utility functions
# ============================================================

def clean_value(value: Any) -> str:
    """
    Convert an Excel value to a trimmed string.

    Empty cells and NaN values become an empty string.
    """

    if value is None or pd.isna(value):
        return ""

    return str(value).strip()


def normalize_column_name(value: Any) -> str:
    """
    Normalize source column names.

    Examples:
        " SP "  -> "sp"
        "Ref"   -> "ref"
        "LABEL" -> "label"
    """

    return clean_value(value).lower()


def normalize_base_ref(value: Any) -> str:
    """
    Normalize a base reference.

    Examples:
        " Y015 "       -> "y015"
        "y015-04/13"   -> "y015"
    """

    ref = clean_value(value).lower()

    if not ref:
        return ""

    # Remove an existing generated suffix if one is present.
    ref = re.sub(r"-\d+/\d+$", "", ref)

    return ref


# ============================================================
# Label splitting
# ============================================================

def split_labels(value: Any) -> list[str]:
    """
    Split a label cell into individual labels.

    Accepted examples:

        潘鋕甲,沈太容,黃亦琴

        潘鋕甲, 沈太容 ,黃亦琴

        潘鋕甲，沈太容，黃亦琴

    Empty items are removed.
    """

    text = clean_value(value)

    if not text:
        return []

    labels = re.split(
        LABEL_SEPARATOR_PATTERN,
        text,
    )

    return [
        label.strip()
        for label in labels
        if label.strip()
    ]


# ============================================================
# Reference generation
# ============================================================

def generate_reference(
    base_ref: str,
    label_index: int,
) -> tuple[str, str]:
    """
    Generate the full reference for one label.

    Returns:
        (generated_reference, warning_message)

    label_index is zero-based.

    Example:
        base_ref = y015
        label_index = 0
        result = y015-04/13
    """

    normalized_ref = normalize_base_ref(base_ref)

    if not normalized_ref:
        return "", "Reference is empty."

    rule = REF_RULES.get(normalized_ref)

    if rule is None:
        # Keep the base reference when no rule exists.
        return (
            normalized_ref,
            f"No reference rule is configured for {normalized_ref}.",
        )

    position = rule.start + (label_index * rule.step)

    generated_ref = (
        f"{normalized_ref}-"
        f"{position:02d}/"
        f"{rule.total:02d}"
    )

    warning = ""

    if position > rule.total:
        warning = (
            f"Generated position {position} exceeds total "
            f"{rule.total} for {normalized_ref}."
        )

    return generated_ref, warning


# ============================================================
# Visual label width and font selection
# ============================================================

def character_visual_width(character: str) -> int:
    """
    Estimate how much horizontal space a character uses.

    Full-width and wide characters, including most Chinese
    characters, count as 2 units.

    English letters and normal punctuation count as 1 unit.
    """

    east_asian_width = unicodedata.east_asian_width(character)

    if east_asian_width in {"W", "F"}:
        return 2

    return 1


def calculate_visual_width(text: str) -> int:
    """
    Calculate the approximate display width of a label.
    """

    return sum(
        character_visual_width(character)
        for character in text
    )


def classify_label_size(
    value: Any,
) -> tuple[str, int, int]:
    """
    Determine the label size category.

    Returns:
        category,
        Word font size,
        calculated visual width
    """

    text = clean_value(value)
    width = calculate_visual_width(text)

    if width <= LARGE_LABEL_MAX_WIDTH:
        return "large", LARGE_FONT_SIZE, width

    if width <= MEDIUM_LABEL_MAX_WIDTH:
        return "medium", MEDIUM_FONT_SIZE, width

    return "small", SMALL_FONT_SIZE, width


def create_word_label_fields(
    value: Any,
) -> dict[str, Any]:
    """
    Put the label into only one of three Word merge fields.

    Word Mail Merge cannot reliably use another Excel column
    as a dynamic font size. Therefore, separate fields are used:

        label_large
        label_medium
        label_small

    Only one field contains text.
    """

    text = clean_value(value)

    category, font_size, visual_width = (
        classify_label_size(text)
    )

    result = {
        "label": text,
        "label_large": "",
        "label_medium": "",
        "label_small": "",
        "label_size": category,
        "label_font_size": font_size,
        "label_visual_width": visual_width,
    }

    if category == "large":
        result["label_large"] = text
    elif category == "medium":
        result["label_medium"] = text
    else:
        result["label_small"] = text

    return result


# ============================================================
# Input validation
# ============================================================

def validate_source_columns(
    source_df: pd.DataFrame,
) -> None:
    """
    Ensure the required source columns exist.
    """

    required_columns = {"sp", "ref", "label"}
    actual_columns = set(source_df.columns)

    missing_columns = required_columns - actual_columns

    if missing_columns:
        missing_text = ", ".join(
            sorted(missing_columns)
        )

        raise ValueError(
            "The input file is missing required columns: "
            f"{missing_text}"
        )


# ============================================================
# Expand raw source records
# ============================================================

def expand_source_rows(
    source_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Expand each source row into individual label records.

    Returns:
        expanded records,
        validation messages
    """

    expanded_rows: list[dict[str, Any]] = []
    validation_rows: list[dict[str, Any]] = []

    for source_index, source_row in source_df.iterrows():
        excel_row_number = source_index + 2

        sp = clean_value(source_row.get("sp"))
        base_ref = normalize_base_ref(
            source_row.get("ref")
        )
        raw_label = clean_value(
            source_row.get("label")
        )

        labels = split_labels(raw_label)

        if not sp:
            validation_rows.append(
                {
                    "source_row": excel_row_number,
                    "severity": "Warning",
                    "sp": sp,
                    "ref": base_ref,
                    "message": "The sp value is empty.",
                }
            )

        if not base_ref:
            validation_rows.append(
                {
                    "source_row": excel_row_number,
                    "severity": "Warning",
                    "sp": sp,
                    "ref": base_ref,
                    "message": "The ref value is empty.",
                }
            )

        if not labels:
            validation_rows.append(
                {
                    "source_row": excel_row_number,
                    "severity": "Warning",
                    "sp": sp,
                    "ref": base_ref,
                    "message": (
                        "No valid labels were found. "
                        "The source row was skipped."
                    ),
                }
            )
            continue

        if base_ref not in REF_RULES:
            validation_rows.append(
                {
                    "source_row": excel_row_number,
                    "severity": "Warning",
                    "sp": sp,
                    "ref": base_ref,
                    "message": (
                        f"No reference rule is configured "
                        f"for {base_ref}."
                    ),
                }
            )

        for label_index, label in enumerate(labels):
            generated_ref, ref_warning = (
                generate_reference(
                    base_ref=base_ref,
                    label_index=label_index,
                )
            )

            label_fields = create_word_label_fields(
                label
            )

            expanded_rows.append(
                {
                    "source_row": excel_row_number,
                    "source_label_index": label_index + 1,
                    "sp": sp,
                    "base_ref": base_ref,
                    "ref": generated_ref,
                    **label_fields,
                }
            )

            if ref_warning:
                validation_rows.append(
                    {
                        "source_row": excel_row_number,
                        "severity": "Warning",
                        "sp": sp,
                        "ref": base_ref,
                        "message": ref_warning,
                    }
                )

    expanded_df = pd.DataFrame(
        expanded_rows,
        columns=[
            "source_row",
            "source_label_index",
            "sp",
            "base_ref",
            "ref",
            "label",
            "label_large",
            "label_medium",
            "label_small",
            "label_size",
            "label_font_size",
            "label_visual_width",
        ],
    )

    validation_df = pd.DataFrame(
        validation_rows,
        columns=[
            "source_row",
            "severity",
            "sp",
            "ref",
            "message",
        ],
    )

    return expanded_df, validation_df


# ============================================================
# Create four-record mail merge format
# ============================================================

def add_blank_record(
    output_row: dict[str, Any],
    record_number: int,
) -> None:
    """
    Add empty fields for an unused label position.
    """

    output_row[f"sp{record_number}"] = ""
    output_row[f"ref{record_number}"] = ""
    output_row[f"label{record_number}"] = ""

    output_row[
        f"label{record_number}_large"
    ] = ""

    output_row[
        f"label{record_number}_medium"
    ] = ""

    output_row[
        f"label{record_number}_small"
    ] = ""

    output_row[
        f"label{record_number}_size"
    ] = ""

    output_row[
        f"label{record_number}_font_size"
    ] = ""

    output_row[
        f"label{record_number}_visual_width"
    ] = ""


def add_record(
    output_row: dict[str, Any],
    record_number: int,
    source_record: pd.Series,
) -> None:
    """
    Add one expanded record to a mail merge row.
    """

    output_row[f"sp{record_number}"] = (
        clean_value(source_record["sp"])
    )

    output_row[f"ref{record_number}"] = (
        clean_value(source_record["ref"])
    )

    output_row[f"label{record_number}"] = (
        clean_value(source_record["label"])
    )

    output_row[
        f"label{record_number}_large"
    ] = clean_value(
        source_record["label_large"]
    )

    output_row[
        f"label{record_number}_medium"
    ] = clean_value(
        source_record["label_medium"]
    )

    output_row[
        f"label{record_number}_small"
    ] = clean_value(
        source_record["label_small"]
    )

    output_row[
        f"label{record_number}_size"
    ] = clean_value(
        source_record["label_size"]
    )

    output_row[
        f"label{record_number}_font_size"
    ] = source_record["label_font_size"]

    output_row[
        f"label{record_number}_visual_width"
    ] = source_record["label_visual_width"]


def create_mail_merge_rows(
    expanded_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Convert expanded records into four-label mail merge rows.
    """

    output_rows: list[dict[str, Any]] = []

    for start_index in range(
        0,
        len(expanded_df),
        GROUP_SIZE,
    ):
        group = expanded_df.iloc[
            start_index:start_index + GROUP_SIZE
        ]

        output_row: dict[str, Any] = {
            "mail_merge_row": (
                len(output_rows) + 1
            )
        }

        for position in range(GROUP_SIZE):
            record_number = position + 1

            if position < len(group):
                source_record = group.iloc[position]

                add_record(
                    output_row=output_row,
                    record_number=record_number,
                    source_record=source_record,
                )
            else:
                add_blank_record(
                    output_row=output_row,
                    record_number=record_number,
                )

        output_rows.append(output_row)

    return pd.DataFrame(output_rows)


# ============================================================
# Workbook formatting
# ============================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

WARNING_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)


def format_worksheet(
    worksheet,
    freeze_cell: str = "A2",
) -> None:
    """
    Apply basic formatting to one worksheet.
    """

    worksheet.freeze_panes = freeze_cell
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.row_dimensions[1].height = 24

    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def set_column_widths(
    worksheet,
    default_width: float = 15,
    maximum_width: float = 45,
) -> None:
    """
    Estimate useful Excel column widths.
    """

    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_index
        )

        maximum_length = 0

        for cell in worksheet[column_letter]:
            cell_value = clean_value(cell.value)

            if not cell_value:
                continue

            line_lengths = [
                calculate_visual_width(line)
                for line in cell_value.splitlines()
            ]

            cell_length = max(
                line_lengths,
                default=0,
            )

            maximum_length = max(
                maximum_length,
                cell_length,
            )

        calculated_width = min(
            max(maximum_length + 2, default_width),
            maximum_width,
        )

        worksheet.column_dimensions[
            column_letter
        ].width = calculated_width


def apply_special_widths(workbook) -> None:
    """
    Apply more suitable widths to important columns.
    """

    if "MailMerge" in workbook.sheetnames:
        worksheet = workbook["MailMerge"]

        for cell in worksheet[1]:
            column_name = clean_value(
                cell.value
            )

            column_letter = get_column_letter(
                cell.column
            )

            if column_name.startswith("sp"):
                worksheet.column_dimensions[
                    column_letter
                ].width = 20

            elif column_name.startswith("ref"):
                worksheet.column_dimensions[
                    column_letter
                ].width = 16

            elif "label" in column_name:
                worksheet.column_dimensions[
                    column_letter
                ].width = 28

    if "Validation" in workbook.sheetnames:
        worksheet = workbook["Validation"]

        for row in worksheet.iter_rows(
            min_row=2,
        ):
            for cell in row:
                cell.fill = WARNING_FILL


def format_output_workbook(
    output_file: Path,
) -> None:
    """
    Open and format the generated workbook.
    """

    workbook = load_workbook(output_file)

    for worksheet in workbook.worksheets:
        format_worksheet(worksheet)
        set_column_widths(worksheet)

    apply_special_widths(workbook)

    workbook.save(output_file)


# ============================================================
# Main conversion process
# ============================================================

def convert_excel(
    input_file: Path,
    output_file: Path,
) -> None:
    """
    Run the complete conversion.
    """

    if not input_file.exists():
        raise FileNotFoundError(
            f"Input file was not found: {input_file}"
        )

    source_df = pd.read_excel(
        input_file,
        sheet_name=0,
        dtype=str,
    ).fillna("")

    source_df.columns = [
        normalize_column_name(column)
        for column in source_df.columns
    ]

    validate_source_columns(source_df)

    expanded_df, validation_df = (
        expand_source_rows(source_df)
    )

    if expanded_df.empty:
        raise ValueError(
            "No valid label records were found."
        )

    mail_merge_df = create_mail_merge_rows(
        expanded_df
    )

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl",
    ) as writer:
        mail_merge_df.to_excel(
            writer,
            sheet_name="MailMerge",
            index=False,
        )

        expanded_df.to_excel(
            writer,
            sheet_name="ExpandedData",
            index=False,
        )

        source_df.to_excel(
            writer,
            sheet_name="OriginalData",
            index=False,
        )

        if validation_df.empty:
            validation_df = pd.DataFrame(
                [
                    {
                        "source_row": "",
                        "severity": "Information",
                        "sp": "",
                        "ref": "",
                        "message": (
                            "No validation warnings "
                            "were found."
                        ),
                    }
                ]
            )

        validation_df.to_excel(
            writer,
            sheet_name="Validation",
            index=False,
        )

    format_output_workbook(output_file)

    print("Conversion completed successfully.")
    print(f"Input file: {input_file.resolve()}")
    print(f"Output file: {output_file.resolve()}")
    print(
        f"Source rows: {len(source_df)}"
    )
    print(
        f"Expanded labels: {len(expanded_df)}"
    )
    print(
        f"Mail merge rows: {len(mail_merge_df)}"
    )
    print(
        f"Validation messages: "
        f"{len(validation_df)}"
    )


def main() -> int:
    """
    Command-line entry point.

    Usage:
        python excel_mail_merge_converter.py

    Or:
        python excel_mail_merge_converter.py source.xlsx result.xlsx
    """

    input_file = INPUT_FILE
    output_file = OUTPUT_FILE

    if len(sys.argv) >= 2:
        input_file = Path(sys.argv[1])

    if len(sys.argv) >= 3:
        output_file = Path(sys.argv[2])

    try:
        convert_excel(
            input_file=input_file,
            output_file=output_file,
        )

        return 0

    except Exception as exc:
        print(
            f"Conversion failed: {exc}",
            file=sys.stderr,
        )

        return 1


if __name__ == "__main__":
    raise SystemExit(main())
